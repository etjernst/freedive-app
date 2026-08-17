"""Single source of truth for the Winnow exercise library.

Defines the canonical exercises (deduped from the two Obsidian library notes)
and the session->canonical mapping, then emits:
  - docs/exercise-canon-and-mapping.xlsx  (human review: catalog + mapping + affinities)
  - seed/fixtures.json                     (the app's seed library, schema-validated)
  - seed/fixtures.seals.json               (the Seals club edition: ten pool exercises with phase defaults)

Run: python seed/build_library.py

Nested sets (DYN 2/17/18): heterogeneous repeated blocks are flattened into
explicit rep rows, with set_repeat for the outer repeat---no recursive group
node. E.g. "2x {50m FRC, 6x25m sprints}" is seven rep rows with set_repeat=2.
"""
import itertools
import json
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = str(ROOT / "docs" / "exercise-canon-and-mapping.xlsx")

# ---------------- canonical catalog ----------------
# Each entry: id, name, discipline, allowed_roles, capacity, fixed_structure, editable_options
catalog = [
    # ----- STA -----
    ("sta-nwu-submax", "NWU sub-max or max", "STA", "main",
     "mental, co2, o2_hypoxia",
     "No warmup; one still sub-max hold to read sensations.",
     "HV prep (def 5 min 5:5); recovery (def full)."),
    ("sta-nwu-submax-moving", "NWU sub-max or max (moving)", "STA", "main",
     "mental, co2, o2_hypoxia",
     "No warmup; one sub-max hold with constant movement (paddle hands, move arms, never still).",
     "HV prep (def 5 min soft 4:6, belly only); recovery (def full)."),
    ("sta-co2-increasing", "CO2 5 x RB increasing", "STA", "main",
     "co2",
     "Fixed recovery breaths; holds increase from a start, until failure.",
     "recovery breaths (def 5); start (def 2:00); step (def +20-30s); termination (def until failure)."),
    ("sta-co2-vshape", "CO2 V-shape", "STA", "main",
     "co2",
     "Holds descend toward ~1 min then ascend; maximize total time under discomfort.",
     "recovery breaths (def 2); start (def ~50% PB); HV prep (def 4:6 soft)."),
    ("sta-co2-1breath", "CO2 1 x RB wonka", "STA", "main",
     "co2",
     "1-breath recovery; short repeated holds until quality drops.",
     "ladder mode: absolute|1C+X (def 1C+X); step (def +10s); termination (def quality-drop)."),
    ("sta-co2-short-intense", "CO2 2 x RB repeats", "STA", "main",
     "co2",
     "5-6 FL holds at a sustainable duration, 2-breath recovery only.",
     "n holds (def 6); hold (def sustainable, e.g. 2:30); recovery (def 2 breaths)."),
    ("sta-co2-decreasing-rec", "CO2 classic", "STA", "main",
     "co2, mental",
     "Fixed hold; recovery shrinks each rep.",
     "hold (def sized to contraction duration); recovery start (def 60s); step (def -10s)."),
    ("sta-co2-square", "Box breathing", "STA", "main",
     "mental, co2",
     "Square breathing held for a long duration.",
     "box length (def 10s); duration (def 30 min)."),
    ("sta-co2-second-hold", "CO2 two-hold", "STA", "main",
     "co2",
     "Sub-max, a few recovery breaths, then a max-effort second hold as the target.",
     "first hold (def ~70% PB); recovery breaths (def 3); prep (def 5 min 5:5)."),
    ("sta-get-high", "Get high (oxygen table)", "STA", "main",
     "o2_hypoxia",
     "Strong HV prep before each; add 1 min per hold until it gets hard, then add 30s only.",
     "HV pattern (def 2:2 strong); start hold (def 1 min); step (def +1 min, then +30s once hard)."),
    ("sta-el", "EL reps", "STA", "main, warmup",
     "o2_hypoxia, mental",
     "RV/EL holds, increasing duration.",
     "n (def 5); style: relaxed-belly|pull-and-release (def relaxed-belly); recovery (def 2-3 min)."),
    ("sta-el-fl-switch", "EL/FL switch", "STA", "main, warmup",
     "o2_hypoxia, mental",
     "Alternating EL and FL holds, both increasing.",
     "EL step (def +20s); FL step (def +20-30s); recovery (def 2 min 4:6 HV); as-warmup: stay below PB margins."),
    ("sta-1c-plus", "First contraction plus", "STA", "main",
     "co2, mental",
     "Increasing breathing prep; hold target = 1C + Xs across reps.",
     "prep ramp (def 2->4 min); 1C+X schedule (def +0/20/40/60/80s); recovery mode (def increasing prep|fixed 2 min)."),
    ("sta-frc-awareness", "FL → FRC", "STA", "main",
     "co2",
     "FL then exhale to FRC then a short hold; short recovery only.",
     "hold steps (def 15/30/45s); recovery (def 1 min)."),
    ("sta-high-volume", "High-volume 70-80% PB", "STA", "main",
     "volume, co2, o2_hypoxia",
     "Repeat a % of PB many times.",
     "% PB (def 70); n reps (def 3-10); rest (def long/unlimited); reduce-prep-each-rep (def off)."),
    ("sta-progressive-fl", "Progressive FL", "STA", "main",
     "co2, mental",
     "Comfortable -> strong sub-max -> max(-ish) FL for the day.",
     "prep (def 5 min); sequence (def 3 holds)."),
    ("sta-max", "Max attempt: STA", "STA", "main",
     "performance",
     "Single maximal breath-hold. The performance test warm-ups are compared against.",
     "HV prep (def 5 min 5:5); lung (def FL)."),
    ("sta-cooldown-easy", "Easy cool-down STA", "STA", "cooldown",
     "mental",
     "Light, relaxed sub-max hold to finish on something easy.",
     "hold (def light submax); recovery (def full)."),
    # STA warm-ups (each its own selectable id)
    ("sta-wu-fl-progression", "WU: FL progression", "STA", "warmup",
     "co2, mental",
     "1C+20 / 1C+40 / 1C+60 FL holds.",
     "steps (def 1C+20/40/60); recovery (def 3 min)."),
    ("sta-wu-frc-progression", "WU: FRC progression", "STA", "warmup",
     "co2",
     "FRC holds through the first hard contractions (not to hypoxia).",
     "n (def 3); recovery (def 3-4 min)."),
    ("sta-wu-contraction-delay", "WU: contraction delay", "STA", "warmup",
     "co2, mental",
     "Repeat holds until 1C arrives at the right time, then go for max.",
     "trigger (def 1C timing feels right)."),
    ("sta-wu-rv-fl", "WU: RV progression + 1 FL", "STA", "warmup",
     "o2_hypoxia, co2",
     "3 RV holds (easy/moderate/hard) + 1 FL.",
     "FL target (def ~1.5 min under PB)."),
    ("sta-wu-nwu-hard-start", "WU: NWU hard start", "STA", "warmup",
     "co2, o2_hypoxia",
     "One hard FL near PB-1min, recover, then max.",
     "hard hold (def PB-1 min); recovery (def 5 min + strong HV)."),
    ("sta-wu-3rv-3fl", "WU: 3xRV + 3xFL", "STA", "warmup",
     "o2_hypoxia",
     "3 EL holds then 3 FL holds.",
     "RV recovery (def 3 min); FL recovery (def 5 min)."),
    # ----- DYN -----
    ("dyn-nwu-submax", "NWU sub-max (dynamic)", "any", "main",
     "mental, o2_hypoxia",
     "~70-80% feel, explore sensations, no fixed distance.",
     "% feel (def 70-80); variant: plain|to-1C+contractions (def plain); discipline (def any; DNF shorter)."),
    ("dyn-max", "Max attempt: dynamic", "any", "main",
     "performance",
     "Single maximal distance attempt. The performance test warm-ups are compared against.",
     "discipline (def any; DNF shorter)."),
    ("dyn-max-simulator", "Max dive simulator", "any", "main",
     "o2_hypoxia",
     "80% PB, short dive, then a max push, with breaths between.",
     "% PB (def 80); short leg (def min{60m,30% PB}); RB (def 2 then 3)."),
    ("dyn-volume-fixed", "Volume: fixed reps", "any", "main",
     "volume, co2",
     "N x a fixed distance with a set recovery (NOT goal-maximize).",
     "n reps; distance (def 50/75m; shorter for DNF); recovery; pacing (def even)."),
    ("dyn-volume-maximize", "Volume: maximize total distance", "any", "main",
     "volume, o2_hypoxia",
     "N dives where the goal is maximizing total distance; distance not fixed.",
     "n dives; recovery (often shrinks across session)."),
    ("dyn-volume-technique", "Volume: technique drill", "DNF", "main",
     "technique, co2",
     "Sets alternating legs-only and normal (or arms-only), minimal recovery.",
     "n sets (def 8-10); block (def 50m legs-only + 50m normal); recovery (def minimal)."),
    ("dyn-pyramid", "Pyramid", "any", "main",
     "co2, volume",
     "25-50-75-50-25; minimal-but-confident recovery; full recovery between sets.",
     "ladder (def 25-50-75-50-25); recovery (def minimal/confident); n pyramids (user adds as many as wanted); distances shorter for DNF."),
    ("dyn-inverse-pyramid", "Pyramid (inverse)", "any", "main",
     "co2, volume",
     "75-50-25-50-75; minimal-but-confident recovery.",
     "ladder (def 75-50-25-50-75); recovery (def < swim time, cap 90s); sets (def 2)."),
    ("dyn-descending", "Descending distance", "any", "main",
     "co2",
     "Long doable distance, then ~20% less, then ~20% less again.",
     "start (def long doable); step (def -20%); n (def 3); recovery (def minimal)."),
    ("dyn-sweet16", "Sweet 16", "any", "main",
     "co2, technique",
     "16 x 25m as quickly as possible.",
     "reps (def 16); distance (def 25m); recovery (def 0-30s)."),
    ("dyn-sprints", "Sprints", "any", "main",
     "fitness_lactic, technique",
     "Max sprints with short recovery.",
     "distance (def 25/50m); n; recovery (short)."),
    ("dyn-frc-sprint", "FRC dive + sprints", "any", "main",
     "fitness_lactic",
     "FRC dive (longest possible) then a max-sprint ladder.",
     "FRC distance (def 50-100m); sprint ladder (def 3x25m, growing recovery); sets (def 3)."),
    ("dyn-elastic-sprint-max", "Elastic sprints into max", "any", "main",
     "fitness_lactic",
     "Elastic-band max sprints then straight into a max dive.",
     "sprints (def 4 x 15s, 15s rest); sets (def 2)."),
    ("dyn-stop-start", "Stop-start (STA then swim)", "any", "main",
     "mental, co2",
     "A STA hold, then straight into a dynamic swim.",
     "STA duration (def 1:30); then distance/sub-max."),
    ("dyn-start-stop", "Start-stop (swim then STA)", "any", "main",
     "mental, co2",
     "A dynamic swim, then straight into a STA hold. (Not in the library; added.)",
     "distance; then STA duration."),
    ("dyn-stop-dyn", "Stop in the middle (dive-STA-dive)", "any", "main",
     "mental, co2",
     "A dynamic, then a STA hold, then another dynamic. (Not in the library; added.)",
     "leg-1 distance; STA duration; leg-2 distance."),
    ("dyn-technique", "Technique drills", "DNF", "main",
     "technique",
     "Arms-only / legs-only / normal, counting strokes to reduce them.",
     "blocks (def 4x each); count (strokes)."),
    ("dyn-tortuga", "Tortuga (slow crawl)", "tortuga", "main",
     "mental, co2",
     "Its own discipline. Cover a short distance in the longest possible time; max time, distance irrelevant.",
     "distance cap (def 50m)."),
    ("dyn-broken-200", "Broken 200m", "any", "main",
     "co2",
     "75m, 50m, 3x25m with minimum recovery; consistent pace.",
     "structure (def 75/50/3x25); recovery (def minimum, caps 60s/30s/few breaths)."),
    ("dyn-hypercapnic", "Hypercapnic cool-down", "any", "cooldown",
     "co2, volume",
     "Continuous swim, breathing every several strokes, no stopping.",
     "distance (def 300-500m); breathe every (def 7+ strokes)."),
]

# ---------------- mapping (one row per source exercise) ----------------
# section, session, source exercise, role used, canonical id, session params (become editable defaults)
mapping = [
    ("STA", "STA 1", "Ex1 NWU sub-max", "main", "sta-nwu-submax", "5 min strong HV (5:5); full recovery"),
    ("STA", "STA 1", "Ex2 short CO2 table", "main", "sta-co2-increasing", "start 2:00, +20-30s, 5xRB, until failure"),
    ("STA", "STA 2", "Ex1 NWU with movement", "main", "sta-nwu-submax-moving", "5 min soft HV (4:6, belly only); full recovery"),
    ("STA", "STA 2", "Ex2 long CO2 table V-shaped", "main", "sta-co2-vshape", "4 min soft HV, start ~50% PB, 2xRB"),
    ("STA", "STA 3", "EL 2x5", "main", "sta-el", "5x relaxed belly + 5x pull-and-release, 2-3 min recovery"),
    ("STA", "STA 4", "first contraction-plus", "main", "sta-1c-plus", "2-4 min prep; 1C +0/20/40/60/80s; alt fixed 2-min recovery"),
    ("STA", "STA 5", "CO2 awareness", "main", "sta-frc-awareness", "FL->FRC, 15/30/45s, 1-min recovery"),
    ("STA", "STA 6", "1-breath table", "main", "sta-co2-1breath", "20-60s FL, 1xRB, moderate urge throughout"),
    ("STA", "STA 7", "get high (O2 table)", "main", "sta-get-high", "2:2 super-strong HV, +30s holds"),
    ("STA", "STA 8", "high-volume", "main", "sta-high-volume", "70% PB x3-10, unlimited rest"),
    ("STA", "STA 8", "(warm-up: 1-3 FRC)", "warmup", "sta-wu-frc-progression", "1-3 FRC holds before the volume work"),
    ("STA", "STA 9", "EL/FL switch", "main", "sta-el-fl-switch", "2 min 4:6 HV recovery; EL +20s / FL +20-30s"),
    ("STA", "STA 10", "high-volume CO2 table", "main", "sta-high-volume", "reduced-prep variant ON; 70% PB x3"),
    ("STA", "STA 10", "(warm-up: 3 RV + 1 FL)", "warmup", "sta-wu-rv-fl", "3 RV holds + 1 long FL"),
    ("STA", "STA 11", "Ex1 6xEL", "main", "sta-el", "6 holds, 2-3 min 5:5 HV"),
    ("STA", "STA 11", "Ex2 1xFL", "main", "sta-progressive-fl", "5 min 5:5 HV, 1 FL at 70-80% PB (single-hold case)"),
    ("STA", "STA 12", "breathing-cycle 1C ladder", "main", "sta-1c-plus", "cycles 10/15/20/25/30; 1C +0/20/40/60 then submax/max"),
    ("STA", "STA 13", "oxygen table", "main", "sta-get-high", "growing prep + holds +20-40s"),
    ("STA", "STA 14", "oxygen table", "main", "sta-get-high", "2 min recovery only, +1 min each, until failure"),
    ("STA", "STA 15", "short & intense CO2", "main", "sta-co2-short-intense", "5-6 FL, 2-breath recovery (e.g. 6x2:30)"),
    ("STA", "STA 15", "(warm-up: 2 FRC)", "warmup", "sta-wu-frc-progression", "2 easy FRC to 4-5 contractions, full recovery"),
    ("STA", "STA 16", "medium & moderate CO2", "main", "sta-co2-decreasing-rec", "hold sized to contraction duration; recovery 60s->10s"),
    ("STA", "STA 16", "(warm-up: 2 FRC)", "warmup", "sta-wu-frc-progression", "2 easy FRC to 4-5 contractions, full recovery"),
    ("STA", "STA 17", "soft & medium CO2", "main", "sta-co2-square", "10s square breathing, 30 min"),
    ("STA", "STA 18", "hypoxia & CO2", "main", "sta-get-high", "2:2 HV, +30s each, target 90s-1min below PB"),
    ("STA", "STA 19", "challenging second hold", "main", "sta-co2-second-hold", "70% PB, 3 RB, max second hold"),
    ("STA", "STA 20", "Ex1 warm-up", "warmup", "sta-wu-frc-progression", "3 FRC progressive + 2 FL 1C+30-60"),
    ("STA", "STA 20", "Ex2 progressive FL", "main", "sta-progressive-fl", "comfortable -> strong sub-max -> max for the day"),
    ("STA", "STA 21", "1-breath CO2 ladder", "main", "sta-co2-1breath", "1C, 1C+10, 1C+20... +10s/rep, until quality drops"),
    ("STA", "STA WU1", "FL progression", "warmup", "sta-wu-fl-progression", "1C+20/40/60, 3 min recovery"),
    ("STA", "STA WU2", "FRC progression", "warmup", "sta-wu-frc-progression", "3 FRC, 3-4 min recovery"),
    ("STA", "STA WU3", "contraction delay", "warmup", "sta-wu-contraction-delay", "repeat to 1C then max once timed right"),
    ("STA", "STA WU4", "RV progression + 1 FL", "warmup", "sta-wu-rv-fl", "3 RV + 1 FL to 1.5 min under PB"),
    ("STA", "STA WU5", "NWU hard start", "warmup", "sta-wu-nwu-hard-start", "hard FL to PB-1min, recover, then max"),
    ("STA", "STA WU6", "3xRV + 3xFL", "warmup", "sta-wu-3rv-3fl", "3 EL (3 min rec) + 3 FL (5 min rec)"),
    ("STA", "STA WU7", "alternating EL/FL", "warmup", "sta-el-fl-switch", "EL/FL below PB margins (same exercise as STA 9, used as warm-up)"),
    # ----- DYN -----
    ("DYN", "DYN 1", "max dive simulator", "main", "dyn-max-simulator", "80% PB, 2xRB, min{60m,30%PB}, 3xRB, max"),
    ("DYN", "DYN 2", "FRC DYNb + sprints", "main", "dyn-frc-sprint", "50-100m FRC, 4xRB, 3x25m sprints 10/20s rec, x3 sets"),
    ("DYN", "DYN 3", "elastic sprint into DNF max", "main", "dyn-elastic-sprint-max", "4x15s elastic, 15s rest, max DNF; 2 sets"),
    ("DYN", "DYN 4", "STA-to-dynamic", "main", "dyn-stop-start", "2-3x STA to 1C/UTB then FL/crawl/FRC, further each dive (was a transition-ladder exercise; folded into stop-start)"),
    ("DYN", "DYN 5", "Ex1 long dives", "main", "dyn-volume-maximize", "3 x long, 3 min recovery, MAXIMIZE total distance"),
    ("DYN", "DYN 5", "Ex2 DNF legs/normal", "main", "dyn-volume-technique", "8-10x {50m legs-only, 50m normal}; DNF"),
    ("DYN", "DYN 6", "Ex1 descending", "main", "dyn-descending", "long, -20%, -20% again, minimal recovery"),
    ("DYN", "DYN 6", "Ex2 sweet 16", "main", "dyn-sweet16", "16 x 25m fast"),
    ("DYN", "DYN 7", "DNF pyramid drills", "main", "dyn-pyramid", "25-50-75-50-25, ~20 min worth of pyramids, 5 min rest, then another set"),
    ("DYN", "DYN 8", "Ex1 stop-start crawl", "main", "dyn-stop-start", "STA then slow dynamic crawl"),
    ("DYN", "DYN 8", "Ex2-4 maximize", "main", "dyn-volume-maximize", "4x/6x/8x any discipline; recovery 2min/1min/30s; MAXIMIZE total distance"),
    ("DYN", "DYN 9", "Ex1 NWU sub-max", "main", "dyn-nwu-submax", "70-80% feel"),
    ("DYN", "DYN 9", "Ex2 technique", "main", "dyn-technique", "4x arms / 4x legs / 4x normal, count strokes (DNF)"),
    ("DYN", "DYN 9", "Ex3 50m DNF (volume)", "main", "dyn-volume-fixed", "4x50m DNF, 1:15-1:30 recovery"),
    ("DYN", "DYN 9", "Ex3 50m DNF (sprint)", "main", "dyn-sprints", "2x50m sprint"),
    ("DYN", "DYN 10", "Ex1 slow crawl (tortuga)", "main", "dyn-tortuga", "50m in max time"),
    ("DYN", "DYN 10", "Ex2 inverse pyramid", "main", "dyn-inverse-pyramid", "2x {75-50-25-50-75} DYNb, recovery < swim time, cap 90s"),
    ("DYN", "DYN 10", "Ex3 50m sprint", "main", "dyn-sprints", "2x50m sprint"),
    ("DYN", "DYN 11", "Ex1 stop-start", "main", "dyn-stop-start", "1:30 STA then 50m+ DNF"),
    ("DYN", "DYN 11", "Ex2 dive plan", "main", "dyn-volume-fixed", "2x75m DNF, 5 min recovery (mental dive-plan focus)"),
    ("DYN", "DYN 11", "Ex3 volume", "main", "dyn-volume-fixed", "3x50m DNF, 1:15 recovery"),
    ("DYN", "DYN 12", "Ex1 NWU sub-max", "main", "dyn-nwu-submax", "70-80% feel"),
    ("DYN", "DYN 12", "Ex2 volume", "main", "dyn-volume-fixed", "6x75m DYNb/DYN, 90-120s recovery (turns/technique focus)"),
    ("DYN", "DYN 12", "Ex3 sprints", "main", "dyn-sprints", "2x50m sprint"),
    ("DYN", "DYN 13", "Ex1 NWU sub-max", "main", "dyn-nwu-submax", "70-80% feel"),
    ("DYN", "DYN 13", "Ex2 inverse pyramid DNF", "main", "dyn-inverse-pyramid", "2x {75-50-25-50-75}, full recovery between sets"),
    ("DYN", "DYN 14", "Ex1 start-stop DNF", "main", "dyn-stop-start", "1:30 STA then DNF sub-max (library labeled 'start-stop'; structurally STA->dive)"),
    ("DYN", "DYN 14", "Ex2 4x75m DNF", "main", "dyn-volume-fixed", "4x75m, 3-5 min recovery"),
    ("DYN", "DYN 14", "Ex3 sprints", "main", "dyn-sprints", "6x25m, 15s recovery"),
    ("DYN", "DYN 14", "500m hypercapnic", "cooldown", "dyn-hypercapnic", "500m hypercapnic swim (listed, no params)"),
    ("DYN", "DYN 15", "Ex1 NWU to 1C+", "main", "dyn-nwu-submax", "variant=to-1C: real discomfort + 3 contractions"),
    ("DYN", "DYN 15", "Ex2 16x25m", "main", "dyn-sweet16", "16x25m, 30s recovery"),
    ("DYN", "DYN 16", "Ex1 NWU sub-max", "main", "dyn-nwu-submax", "75%+ feel"),
    ("DYN", "DYN 16", "Ex2 stop-start", "main", "dyn-stop-start", "2x {1 min STA + 50m+ DNF}, full recovery"),
    ("DYN", "DYN 16", "Ex3 volume", "main", "dyn-volume-fixed", "6x50m DNF, 75-90s recovery"),
    ("DYN", "DYN 17", "Ex1 broken 200m DYNb", "main", "dyn-broken-200", "75/50/3x25, minimum recovery, consistent pace"),
    ("DYN", "DYN 17", "Ex2 sprints", "main", "dyn-sprints", "4x50m max sprint, max 2 min recovery"),
    ("DYN", "DYN 18", "Ex1 FRC + sprints", "main", "dyn-frc-sprint", "2x {50m FRC, 4xRB, 6x25m sprints, 15s recovery}"),
    ("DYN", "DYN 18", "Ex2 paced 75s", "main", "dyn-volume-fixed", "6x75m, recovery < 2 min, fast first 25m, sprint last 50m"),
    ("DYN", "Cool-down", "hypercapnic", "cooldown", "dyn-hypercapnic", "300-500m, breathe every 7+ strokes, no stopping"),
]

cat_by_id = {c[0]: c for c in catalog}

# sanity: every mapping canonical id must exist
missing = sorted({m[4] for m in mapping} - set(cat_by_id))
assert not missing, f"mapping refers to unknown ids: {missing}"

# ---------------- affinities (co-occurrence within a source session) ----------------
session_ids = defaultdict(list)
for (section, session, ex, role, cid, params) in mapping:
    if cid not in session_ids[session]:
        session_ids[session].append(cid)
pair_sessions = defaultdict(list)
for session, ids in session_ids.items():
    for a, b in itertools.combinations(sorted(ids), 2):
        pair_sessions[(a, b)].append(session)
affinities = sorted(pair_sessions.items(), key=lambda kv: (-len(kv[1]), kv[0]))

# ---------------- workbook ----------------
wb = openpyxl.Workbook()

ws2 = wb.active
ws2.title = "Canonical catalog"
ws2.append(["Canonical id", "Name", "Discipline", "Allowed roles", "Capacity focus",
            "Fixed structure", "Editable options (defaults)"])
for c in catalog:
    ws2.append(list(c))

ws = wb.create_sheet("Mapping (combined)")
ws.append(["Section", "Session", "Source exercise", "Role used", "Canonical id",
           "Canonical name", "Discipline", "Allowed roles", "Capacity focus",
           "Fixed structure", "Session params (-> editable defaults)"])
for (section, session, ex, role, cid, params) in mapping:
    c = cat_by_id[cid]
    ws.append([section, session, ex, role, cid, c[1], c[2], c[3], c[4], c[5], params])

ws3 = wb.create_sheet("Affinities")
ws3.append(["Exercise A", "Exercise B", "# sessions together", "Sessions"])
for (a, b), sessions in affinities:
    ws3.append([a, b, len(sessions), ", ".join(sessions)])

# ---------------- styling ----------------
hdr_fill = PatternFill("solid", fgColor="4F6138")
hdr_font = Font(color="FFFFFF", bold=True, size=11)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
warm_fill = PatternFill("solid", fgColor="EAF1DC")

def style_sheet(ws, widths, role_col=None):
    ws.freeze_panes = "A2"
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        is_wu = role_col is not None and str(row[role_col].value) == "warmup"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if is_wu:
                cell.fill = warm_fill
    ws.auto_filter.ref = ws.dimensions

style_sheet(ws2, [24, 30, 10, 13, 22, 46, 52])
style_sheet(ws, [8, 11, 24, 9, 22, 26, 10, 13, 20, 40, 44], role_col=3)
style_sheet(ws3, [24, 24, 12, 40])

# ---------------- seed/fixtures.json (single source of truth) ----------------
FIXTURES = str(ROOT / "seed" / "fixtures.json")
FIXTURES_SEALS = str(ROOT / "seed" / "fixtures.seals.json")

# primary role per template (allowed_roles is a higher-level concept the
# Phase 0a schema does not model; the seed carries the primary role).
WARMUP_IDS = {c[0] for c in catalog if c[3] == "warmup"}
COOLDOWN_IDS = {"sta-cooldown-easy", "dyn-hypercapnic"}

def role_for(cid):
    if cid in WARMUP_IDS:
        return "warmup"
    if cid in COOLDOWN_IDS:
        return "cooldown"
    return "main"

SHAPE = {
    "dyn-stop-start": "stop-start",
    "dyn-start-stop": "start-stop",
    "dyn-stop-dyn": "stop-in-the-middle",
    "dyn-hypercapnic": "continuous-protocol",
    "sta-co2-square": "continuous-protocol",
}
ENV = {"sta-co2-square": "dry"}  # everything else defaults to pool

# Termination describes only how a set ENDS. A fixed_n carries NO count: the
# planned count is reps.length x set_repeat (see plannedRepCount in the app). So
# only the genuinely open-ended / capped / range cases are listed here; every
# other exercise defaults to {"type": "fixed_n"} with the count in its reps.
TERMINATION = {
    "sta-co2-increasing": {"type": "until_failure"},
    "sta-get-high": {"type": "until_failure"},
    "sta-el-fl-switch": {"type": "until_failure"},
    "sta-co2-1breath": {"type": "until_quality_drops"},
    "sta-co2-square": {"type": "duration_capped", "duration_s": 1800},
}

# Outer-repeat count (default 1 elsewhere): flattened nested sets, plus the
# "identical reps" tables encoded as one rep row repeated N times rather than as
# N listed rows. Total planned reps is always reps.length x set_repeat.
_BREATHS = lambda n: {"type": "absolute", "value": n, "unit": "breaths"}
_SECS = lambda n: {"type": "absolute", "value": n, "unit": "time"}
_MINIMAL = {"type": "qualitative", "value": "minimal"}
_ADEQUATE = {"type": "qualitative", "value": "adequate"}
_FULL = {"type": "qualitative", "value": "full"}

SET_REPEAT = {
    "dyn-frc-sprint": 2,
    "sta-co2-short-intense": 6,      # 6 identical sustainable FL holds, 2-breath rec
    "dyn-sweet16": 16,               # 16 x 25m sprints
    "dyn-volume-technique": 8,       # 8 x {50m legs-only + 50m normal}
    "dyn-elastic-sprint-max": 2,     # 2 sets of {4 elastic sprints + max DNF}
}

# Between-sets recovery for the set_repeat exercises whose sets are real blocks
# (not the identical-reps encoding trick, where the per-rep recovery is the rest).
RECOVERY_INTER = {
    "dyn-frc-sprint": _SECS(240),          # up to 4 min between sets (DYN 2)
    "dyn-elastic-sprint-max": _SECS(180),  # 3 min between sets (DYN 3)
    "dyn-volume-technique": _SECS(150),    # 2-3 min between sets (DYN 5)
    "dyn-pyramid": _SECS(300),             # 5 min rest, then another set (DYN 7)
    "dyn-inverse-pyramid": _ADEQUATE,      # adequate recovery between sets (DYN 10)
}

# Pre-built rep rows for the compound / nested-set exercises, so a fresh install
# (or a device after "Refresh library") gets the real structure instead of one
# placeholder rep. Everything else defaults to a single editable rep below.
REPS = {
    # DYN 1: 80% PB, short leg (min{60m,30% PB}), then a max push.
    "dyn-max-simulator": [
        {"shape": "simple", "distance_target": {"unit": "pct_pb", "value": 80},
         "recovery": _BREATHS(2), "note": "80% PB dive"},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 60},
         "recovery": _BREATHS(3), "note": "short leg (min{60m, 30% PB})"},
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "max"},
         "note": "max push"},
    ],
    # DYN 17 Ex1: broken 200m, 75/50/3x25, minimum recovery, consistent pace.
    "dyn-broken-200": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 75}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 25}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 25}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 25}},
    ],
    # DYN 2 / DYN 18 Ex1: 2x {50m FRC, 4 breaths, 6x25m sprints, 15s recovery}.
    "dyn-frc-sprint": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "recovery": _BREATHS(4), "lung_volume": "FRC", "note": "FRC"},
        *[{"shape": "simple", "distance_target": {"unit": "absolute", "value": 25},
           "recovery": _SECS(15), "pace": "sprint", "note": "sprint"} for _ in range(6)],
    ],
    # STA 15: 6 identical sustainable FL holds (e.g. 2:30), 2-breath recovery.
    # One rep row repeated 6x via SET_REPEAT.
    "sta-co2-short-intense": [
        {"shape": "simple", "hold_target": {"unit": "absolute", "value": 150},
         "recovery": _BREATHS(2)},
    ],
    # STA 2: V-shaped CO2 table, holds descend 2:30 -> 1:00 then climb back up,
    # 2-breath recovery throughout.
    "sta-co2-vshape": [
        {"shape": "simple", "hold_target": {"unit": "absolute", "value": s},
         "recovery": _BREATHS(2)}
        for s in (150, 135, 120, 105, 90, 60, 60, 90, 105, 120, 135, 150)
    ],
    # STA 16: fixed sub-max hold, recovery shrinks 60s -> 10s across 7 reps.
    "sta-co2-decreasing-rec": [
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "submax"},
         "recovery": _SECS(r)} for r in (60, 50, 40, 30, 20, 15, 10)
    ],
    # STA 4 / STA 12: 1C+X ladder, +20s per rep, fixed ~2 min recovery.
    "sta-1c-plus": [
        {"shape": "simple", "hold_target": {"unit": "contraction_relative", "value": x},
         "recovery": _SECS(120)} for x in (0, 20, 40, 60, 80)
    ],
    # STA 5: each step is a 1 min FL hold, exhale to FRC with no breath (0s
    # recovery), then an FRC hold of 15/30/45s. Both holds count toward time
    # under breath-hold; 1 min recovery between steps builds difficulty.
    "sta-frc-awareness": [
        rep
        for s in (15, 30, 45)
        for rep in (
            {"shape": "simple", "hold_target": {"unit": "absolute", "value": 60},
             "lung_volume": "FL", "recovery": _SECS(0), "note": "exhale to FRC, no breath"},
            {"shape": "simple", "hold_target": {"unit": "absolute", "value": s},
             "lung_volume": "FRC", "recovery": _SECS(60)},
        )
    ],
    # STA 20: comfortable -> strong sub-max -> max-ish FL, full recovery between.
    "sta-progressive-fl": [
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "submax"}, "recovery": _FULL},
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "strong_submax"}, "recovery": _FULL},
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "max"}},
    ],
    # DYN 6 Ex1: long doable distance, then ~20% less, then ~20% less again.
    "dyn-descending": [
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "long (doable)"}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "~20% less"}, "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "~20% less again"}},
    ],
    # DYN 7: 25-50-75-50-25 pyramid, minimal-but-confident recovery. Add reps to
    # extend; full recovery between sets.
    "dyn-pyramid": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": d}, "recovery": _MINIMAL}
        for d in (25, 50, 75, 50, 25)
    ],
    # DYN 10 / DYN 13: 75-50-25-50-75 inverse pyramid, recovery < swim time, cap 90s.
    "dyn-inverse-pyramid": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": d},
         "recovery": {"type": "cap", "value": 90, "unit": "time"}}
        for d in (75, 50, 25, 50, 75)
    ],
    # DYN 6 Ex2 / DYN 15: 16 x 25m sprints, minimal recovery. One rep repeated 16x.
    "dyn-sweet16": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 25},
         "pace": "sprint", "recovery": _MINIMAL},
    ],
    # ---- first-pass structure for the previously bare templates ----
    # Example values grounded in the library notes and the logged history
    # (2026-01 to 2026-06); flagged inline where invented.
    # STA 1 Ex1: 5 min strong HV, one sub-max hold, full recovery.
    "sta-nwu-submax": [
        {"shape": "simple", "prep_breathing": {"pattern": "5:5", "duration_s": 300},
         "hold_target": {"unit": "qualitative", "value": "submax"}},
    ],
    # STA 2 Ex1: soft HV, sub-max with constant movement.
    "sta-nwu-submax-moving": [
        {"shape": "simple", "prep_breathing": {"pattern": "4:6", "duration_s": 300},
         "hold_target": {"unit": "qualitative", "value": "submax"},
         "note": "constant movement: paddle hands, move arms, never still"},
    ],
    # STA 1 Ex2: start 2:00, +20s per hold, 5 recovery breaths, until failure.
    "sta-co2-increasing": [
        {"shape": "simple", "hold_target": {"unit": "absolute", "value": s},
         "recovery": _BREATHS(5)} for s in (120, 140, 160, 180, 200, 220)
    ],
    # STA 21: 1C ladder, +10s per rep, 1-breath recovery, until quality drops.
    "sta-co2-1breath": [
        {"shape": "simple", "hold_target": {"unit": "contraction_relative", "value": x},
         "recovery": _BREATHS(1)} for x in (0, 10, 20, 30, 40, 50)
    ],
    # STA 17: 10s square breathing for 30 minutes (dry).
    "sta-co2-square": [
        {"shape": "continuous-protocol",
         "continuous": {"duration_s": 1800,
                        "pattern": "10s square (10 in / 10 hold / 10 out / 10 hold)"}},
    ],
    # STA 19: ~70% PB, 3 recovery breaths, then the max-effort second hold.
    "sta-co2-second-hold": [
        {"shape": "simple", "prep_breathing": {"pattern": "5:5", "duration_s": 300},
         "hold_target": {"unit": "pct_pb", "value": 70}, "recovery": _BREATHS(3)},
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "max"},
         "note": "the target: aim to increase this second hold"},
    ],
    # STA 7/13/14 oxygen table; ladder from the 2026-02-22 log (1:00 -> 4:30,
    # 1 min 2:2 HV between), +1 min per hold then +30s once hard, until failure.
    "sta-get-high": [
        {"shape": "simple", "prep_breathing": {"pattern": "2:2", "duration_s": 60},
         "hold_target": {"unit": "absolute", "value": s}, "recovery": _SECS(60)}
        for s in (60, 120, 180, 240, 270, 300)
    ],
    # STA 3/11: EL holds increasing progressively, 2-3 min recovery.
    # Ladder invented around the logged EL range (0:45-2:00; EL max 2:42).
    "sta-el": [
        {"shape": "simple", "lung_volume": "RV",
         "hold_target": {"unit": "absolute", "value": s}, "recovery": _SECS(150)}
        for s in (60, 75, 90, 105, 120)
    ],
    # STA 9: alternating EL/FL ladder as logged 2026-03-19 (EL +15s, FL +30s),
    # 2 min 4:6 HV recovery, until failure.
    "sta-el-fl-switch": [
        rep
        for el, fl in ((45, 150), (60, 180), (75, 210), (90, 240))
        for rep in (
            {"shape": "simple", "lung_volume": "RV",
             "hold_target": {"unit": "absolute", "value": el},
             "recovery": _SECS(120), "note": "4:6 HV during recovery"},
            {"shape": "simple", "lung_volume": "FL",
             "hold_target": {"unit": "absolute", "value": fl},
             "recovery": _SECS(120)},
        )
    ],
    # STA 8: 70% PB repeated (def 3x), unlimited rest.
    "sta-high-volume": [
        {"shape": "simple", "hold_target": {"unit": "pct_pb", "value": 70},
         "recovery": _FULL} for _ in range(3)
    ],
    # STA 1 max attempt: 5 min 5:5 prep, one max hold.
    "sta-max": [
        {"shape": "simple", "prep_breathing": {"pattern": "5:5", "duration_s": 300},
         "hold_target": {"unit": "qualitative", "value": "max"}},
    ],
    # Light, relaxed sub-max to finish on something easy.
    "sta-cooldown-easy": [
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "submax"},
         "note": "light and relaxed; finish on something easy"},
    ],
    # WU1: 1C+20 / 1C+40 / 1C+60, 3 min recovery.
    "sta-wu-fl-progression": [
        {"shape": "simple", "hold_target": {"unit": "contraction_relative", "value": x},
         "recovery": _SECS(180)} for x in (20, 40, 60)
    ],
    # WU2: 3 FRC holds through the first hard contractions, 3-4 min recovery.
    # Durations from the 2026-06-21 log (1:30 / 1:45 / 2:00).
    "sta-wu-frc-progression": [
        {"shape": "simple", "lung_volume": "FRC",
         "hold_target": {"unit": "absolute", "value": s}, "recovery": _SECS(210),
         **({"note": "through the first hard contractions, not to hypoxia"} if s == 90 else {})}
        for s in (90, 105, 120)
    ],
    # WU3: repeat holds to 1C; go for max once 1C arrives at the right time.
    "sta-wu-contraction-delay": [
        {"shape": "simple", "hold_target": {"unit": "contraction_relative", "value": 0},
         "recovery": _FULL, "note": "repeat until 1C arrives at the right time"},
        {"shape": "simple", "hold_target": {"unit": "contraction_relative", "value": 0},
         "recovery": _FULL},
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "max"},
         "note": "only once the 1C timing feels right"},
    ],
    # WU4: 3 RV holds (easy/moderate/hard) + 1 FL to ~1.5 min under PB.
    "sta-wu-rv-fl": [
        {"shape": "simple", "lung_volume": "RV",
         "hold_target": {"unit": "qualitative", "value": "first_discomfort"},
         "recovery": _SECS(180), "note": "easy"},
        {"shape": "simple", "lung_volume": "RV",
         "hold_target": {"unit": "qualitative", "value": "submax"},
         "recovery": _SECS(180), "note": "moderate"},
        {"shape": "simple", "lung_volume": "RV",
         "hold_target": {"unit": "qualitative", "value": "strong_submax"},
         "recovery": _SECS(180), "note": "hard"},
        {"shape": "simple", "lung_volume": "FL",
         "hold_target": {"unit": "absolute", "value": 240},
         "note": "~1.5 min under PB; at least 1 min of contractions"},
    ],
    # WU5: hard FL to ~PB-1 min, 5 min recovery with strong HV, then max.
    "sta-wu-nwu-hard-start": [
        {"shape": "simple", "prep_breathing": {"pattern": "5:5", "duration_s": 300},
         "hold_target": {"unit": "absolute", "value": 270},
         "recovery": _SECS(300), "note": "~PB-1 min; strong HV during recovery"},
        {"shape": "simple", "hold_target": {"unit": "qualitative", "value": "max"}},
    ],
    # WU6: 3 EL (3 min rec) + 3 FL (5 min rec). Durations invented around the
    # logged EL/FL ladders.
    "sta-wu-3rv-3fl": [
        *({"shape": "simple", "lung_volume": "RV",
           "hold_target": {"unit": "absolute", "value": s}, "recovery": _SECS(180)}
          for s in (60, 75, 90)),
        *({"shape": "simple", "lung_volume": "FL",
           "hold_target": {"unit": "absolute", "value": s}, "recovery": _SECS(300)}
          for s in (150, 180, 210)),
    ],
    # DYN 9/12/13 Ex1: ~70-80% feel, no fixed distance.
    "dyn-nwu-submax": [
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "70-80% feel"},
         "note": "no distance in mind; focus on the sensations"},
    ],
    # Single maximal attempt.
    "dyn-max": [
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "max"}},
    ],
    # DYN 9 Ex3: 4 x 50m DNF, 1:15-1:30 recovery.
    "dyn-volume-fixed": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "recovery": _SECS(75)} for _ in range(4)
    ],
    # DYN 8 Ex2: N dives, fixed recovery, maximize total distance.
    "dyn-volume-maximize": [
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "maximize"},
         "recovery": _SECS(120),
         **({"note": "goal: maximize total distance across the dives"} if i == 0 else {})}
        for i in range(4)
    ],
    # DYN 5 Ex2: 8 x {50m legs-only + 50m normal}, minimal recovery in the set.
    "dyn-volume-technique": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "technique_variant": "legs_only", "recovery": _MINIMAL},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "technique_variant": "normal"},
    ],
    # DYN 17 Ex2: max sprints, capped recovery (def 2 x 50m).
    "dyn-sprints": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "pace": "max_sprint", "recovery": {"type": "cap", "value": 120, "unit": "time"}}
        for _ in range(2)
    ],
    # DYN 3: 4 x 15s elastic-band max sprints (15s rest), straight into a max
    # DNF; 2 sets. The sprints are timed surface efforts, encoded as
    # qualitative distances.
    "dyn-elastic-sprint-max": [
        *({"shape": "simple",
           "distance_target": {"unit": "qualitative", "value": "15s elastic sprint"},
           "pace": "max_sprint", "recovery": _SECS(15)} for _ in range(3)),
        {"shape": "simple",
         "distance_target": {"unit": "qualitative", "value": "15s elastic sprint"},
         "pace": "max_sprint", "recovery": _SECS(0), "note": "straight into the max"},
        {"shape": "simple", "distance_target": {"unit": "qualitative", "value": "max"},
         "note": "max DNF immediately after the last sprint"},
    ],
    # DYN 11/16 + 2026-06-01 log: 1:30 STA then 50m+ DNF, further each rep.
    "dyn-stop-start": [
        {"shape": "stop-start", "hold_target": {"unit": "absolute", "value": 90},
         "distance_target": {"unit": "absolute", "value": 50}, "recovery": _FULL,
         "note": "swim further on each rep"},
        {"shape": "stop-start", "hold_target": {"unit": "absolute", "value": 90},
         "distance_target": {"unit": "absolute", "value": 65}},
    ],
    # 2026-06-01 log: DNF to ~60m then a short STA, longer hold each rep.
    "dyn-start-stop": [
        {"shape": "start-stop", "distance_target": {"unit": "absolute", "value": 60},
         "hold_target": {"unit": "absolute", "value": 15}, "recovery": _FULL,
         "note": "hold longer on each rep"},
        {"shape": "start-stop", "distance_target": {"unit": "absolute", "value": 60},
         "hold_target": {"unit": "absolute", "value": 20}},
    ],
    # Dive, mid-pool STA, dive again. Example values invented (no log yet).
    "dyn-stop-dyn": [
        {"shape": "stop-in-the-middle",
         "distance_target": {"unit": "absolute", "value": 50},
         "hold_target": {"unit": "absolute", "value": 20},
         "distance2_target": {"unit": "absolute", "value": 25}, "recovery": _FULL},
        {"shape": "stop-in-the-middle",
         "distance_target": {"unit": "absolute", "value": 50},
         "hold_target": {"unit": "absolute", "value": 30},
         "distance2_target": {"unit": "absolute", "value": 25}},
    ],
    # DYN 9 Ex2: 4x arms-only / 4x legs-only / 4x normal, counting strokes.
    "dyn-technique": [
        *({"shape": "simple", "distance_target": {"unit": "absolute", "value": 25},
           "technique_variant": "arms_only", "recovery": _ADEQUATE,
           **({"note": "count strokes; aim to reduce them"} if i == 0 else {})}
          for i in range(4)),
        *({"shape": "simple", "distance_target": {"unit": "absolute", "value": 25},
           "technique_variant": "legs_only", "recovery": _ADEQUATE} for _ in range(4)),
        *({"shape": "simple", "distance_target": {"unit": "absolute", "value": 25},
           "technique_variant": "normal", "recovery": _ADEQUATE} for _ in range(4)),
    ],
    # DYN 10 Ex1: <=50m in the longest possible time (2 reps logged 2026-04-30).
    "dyn-tortuga": [
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "pace": "relaxed", "recovery": _ADEQUATE,
         "note": "longest possible time; distance irrelevant"},
        {"shape": "simple", "distance_target": {"unit": "absolute", "value": 50},
         "pace": "relaxed"},
    ],
    # Cool-down: 300-500m continuous hypercapnic swim, breathe every 7+ strokes.
    "dyn-hypercapnic": [
        {"shape": "continuous-protocol",
         "distance_target": {"unit": "absolute", "value": 400},
         "continuous": {"stroke_cadence": "every 7+ strokes"},
         "note": "300-500m, no stopping"},
    ],
}

# Uniform lung volume / speed stamped onto placeholder reps (applied to every
# rep) for exercises NOT in REPS. Every current exercise carries its structure
# in REPS, so these are empty; they remain for future placeholder templates.
# lung_volume uses the schema codes FL/FRC/RV (RV is empty lung / EL).
LUNG = {}
SPEED = {}
# Exercises logged as a summary (lap distance + total time + rep count) rather
# than rep by rep. Everything else defaults to per_rep.
LOG_MODE = {
    "dyn-sweet16": "aggregate",
}


# ---------------- Seals edition ----------------
# The club edition ships these ten pool exercises and nothing else. Same tuple
# shape as the canonical catalog; every dynamic one is DYNb, tortuga stays its
# own discipline. Numbers are the design proposals in docs/seals-basic-program.md
# (Jeranko's sequencing at club intensity), not values read off his sheets.
seals_catalog = [
    ("seals-fixed-rest", "Fixed-rest set", "DYNb", "main",
     "co2",
     "Repeated 50 m at your normal dive pace with a rest cap. Base: 4, 5, 6, 7 reps across the four weeks, cap 60 s. Specialization: 6 reps, cap 45 s. Taper: 5 then 4 reps, cap 45 s.",
     "reps (def 4); distance (def 50m, 25m for newer members); rest cap (def 60s)."),
    ("seals-pyramid", "Pyramid", "DYNb", "main",
     "co2",
     "50-25-25-50 in weeks 1-2, 50-50-25-50-50 in weeks 3-4, rest no longer than 60 s and shorter when you can.",
     "ladder (def 50-25-25-50); rest cap (def 60s)."),
    ("seals-hypercapnic-swim", "Hypercapnic surface swim", "any", "main",
     "co2",
     "Freestyle on the surface in short fins, no wetsuit, breathing a fixed 3 times per 25 m, no stopping. Base 100 to 250 m, build 300 to 350 m, specialization 400 m, taper 300 then 200 m.",
     "distance (def 100m); breaths per 25m (def 3)."),
    ("seals-recovery-swim", "Recovery swim", "any", "cooldown",
     "",
     "100 m any style, relaxed, breathing normally, to close every session.",
     "distance (def 100m)."),
    ("seals-over-under", "Over-under warm-up", "DYNb", "warmup",
     "technique",
     "25 m freestyle on the surface, turn, 25 m DYNb underwater without stopping; twice for 100 m.",
     "rounds (def 2)."),
    ("seals-sprints", "Sprints", "DYNb", "main",
     "fitness_lactic",
     "All-out DYNb sprints. Base: 2 sets of 6 x 25 m, 20 s between reps, 2 min between sets (weeks 1-2), then 6 x 50 m with 60 s (weeks 3-4). Build: 5-6 x 75 m with 75-120 s.",
     "distance (def 25m); reps (def 6); sets (def 2); rest (def 20s); rest between sets (def 2 min)."),
    ("seals-dolphin-sprints", "Dolphin sprints", "DYNb", "main",
     "fitness_lactic",
     "Each rep is 25 m of all-out dolphin kick on the surface (snorkel, or on your back so you can breathe), then a duck dive straight into 10 m underwater at sprint pace. 2 sets of 4 (weeks 5-6), 2 sets of 5 (weeks 7-8), 60-90 s between reps, 3 min between sets, short fins, no wetsuit.",
     "reps (def 4); sets (def 2); rest (def 90s); rest between sets (def 3 min)."),
    ("seals-submax", "Submax dive", "DYNb", "main",
     "o2_hypoxia, mental",
     "One dive, no warm-up, no breathe-up, breathing normally beforehand, face dry until you go. Aim for at least the target percentage of your PB with no upper limit, but surface before your limit and never finish shaky. Build 65%, specialization 70-75%, taper 80%.",
     "% of PB (def 65)."),
    ("seals-endurance-ladder", "Endurance ladder", "DYNb", "main",
     "co2, volume",
     "Mixed 75 m and 50 m with rest no longer than the dive time. Week 9: 1 x 75 + 5 x 50, then convert one 50 into a 75 each week (2 + 4, 3 + 3, 4 + 2).",
     "75s (def 1); 50s (def 5); rest (def at most the dive time)."),
    ("seals-tortuga", "Tortuga (slow crawl)", "tortuga", "main",
     "mental, co2",
     "Crawl along the bottom as slowly as you can, covering at most 50 m in the longest possible time; no goggles so you do not watch the clock. From week 5, one slow crawl after the sprints, every week.",
     "distance cap (def 50m)."),
]

_CAP = lambda n: {"type": "cap", "value": n, "unit": "time"}
_LE_SWIM = {"type": "inequality", "value": "<= swim time"}
_D = lambda m, **kw: {"shape": "simple", "distance_target": {"unit": "absolute", "value": m}, **kw}
_PCT = lambda p, **kw: {"shape": "simple", "distance_target": {"unit": "pct_pb", "value": p}, **kw}
_SWIM = lambda m, note=None: {"shape": "continuous-protocol",
                              "distance_target": {"unit": "absolute", "value": m},
                              "continuous": {"stroke_cadence": "3 breaths per 25 m"},
                              **({"note": note} if note else {})}

def _fixed_rest(n, cap):
    return [_D(50, recovery=_CAP(cap)) for _ in range(n)]

def _dolphin(n, rest=90):
    rows = []
    for i in range(n):
        rows.append(_D(25, pace="max_sprint", note="surface dolphin kick, snorkel or on your back"))
        rows.append(_D(10, pace="max_sprint", recovery=_CAP(rest), note="duck dive, underwater sprint"))
    return rows

def _ladder(n75, n50):
    return [_D(75, recovery=_LE_SWIM) for _ in range(n75)] + [_D(50, recovery=_LE_SWIM) for _ in range(n50)]

_RELAXED_100 = lambda: [_D(100, pace="relaxed", note="any style, breathe normally")]
_OVER_UNDER = lambda: [_D(25, note="freestyle on the surface"),
                       _D(25, note="turn, DYNb underwater, no stop")]
_TORTUGA = lambda: [_D(50, pace="relaxed", note="longest possible time; distance irrelevant")]

# Per-phase defaults: reps (and set_repeat / recovery_inter where they change)
# that the builder applies when an exercise is added under a chosen phase. The
# template's own reps are the first phase's defaults, so an app without phases
# still gets a sensible plan.
SEALS_PHASES = {
    "seals-fixed-rest": {
        "base": {"reps": _fixed_rest(4, 60)},
        "specialization": {"reps": _fixed_rest(6, 45)},
        "taper": {"reps": _fixed_rest(5, 45)},
    },
    "seals-pyramid": {
        "base": {"reps": [_D(50, recovery=_CAP(60)), _D(25, recovery=_CAP(60)),
                          _D(25, recovery=_CAP(60)), _D(50)]},
    },
    "seals-hypercapnic-swim": {
        "base": {"reps": [_SWIM(100, "100 to 250 m across the block")]},
        "build": {"reps": [_SWIM(300, "300 to 350 m across the block")]},
        "specialization": {"reps": [_SWIM(400)]},
        "taper": {"reps": [_SWIM(300, "300 then 200 m")]},
    },
    "seals-recovery-swim": {p: {"reps": _RELAXED_100()} for p in ("base", "build", "specialization", "taper")},
    "seals-over-under": {p: {"reps": _OVER_UNDER(), "set_repeat": 2} for p in ("base", "build")},
    "seals-sprints": {
        "base": {"reps": [_D(25, pace="max_sprint", recovery=_SECS(20)) for _ in range(6)],
                 "set_repeat": 2, "recovery_inter": _SECS(120)},
        "build": {"reps": [_D(75, pace="max_sprint", recovery=_CAP(120)) for _ in range(5)],
                  "set_repeat": 1},
    },
    "seals-dolphin-sprints": {
        "build": {"reps": _dolphin(4), "set_repeat": 2, "recovery_inter": _SECS(180)},
    },
    "seals-submax": {
        "build": {"reps": [_PCT(65, note="no upper limit; surface before your limit")]},
        "specialization": {"reps": [_PCT(70, note="70-75%; no upper limit; surface before your limit")]},
        "taper": {"reps": [_PCT(80, note="no upper limit; surface before your limit")]},
    },
    "seals-endurance-ladder": {
        "specialization": {"reps": _ladder(1, 5)},
        "taper": {"reps": _ladder(3, 3)},
    },
    "seals-tortuga": {p: {"reps": _TORTUGA()} for p in ("build", "specialization")},
}
PHASE_ORDER = ["base", "build", "specialization", "taper"]

seals_templates = []
for (cid, name, discipline, role, capacity, structure, options) in seals_catalog:
    phases = SEALS_PHASES[cid]
    first = next(p for p in PHASE_ORDER if p in phases)
    d0 = phases[first]
    seals_templates.append({
        "schema_version": 1,
        "id": cid,
        "name": name,
        "environment": "pool",
        "role": role,
        "discipline": discipline,
        "capacity_tags": [c.strip() for c in capacity.split(",") if c.strip()],
        "collections": ["seals"],
        "phase_tags": [p for p in PHASE_ORDER if p in phases],
        "phase_defaults": phases,
        "goal": structure,
        "cues": options,
        "log_mode": "per_rep",
        "set_repeat": d0.get("set_repeat", 1),
        **({"recovery_inter": d0["recovery_inter"]} if "recovery_inter" in d0 else {}),
        "termination": {"type": "fixed_n"},
        "reps": d0["reps"],
    })

ws_seals = wb.create_sheet("Seals catalog")
ws_seals.append(["Id", "Name", "Discipline", "Role", "Capacity focus", "Structure", "Editable options (defaults)"])
for c in seals_catalog:
    ws_seals.append(list(c))
style_sheet(ws_seals, [24, 26, 10, 10, 18, 60, 44], role_col=3)
# The workbook is review output; if Excel has it open, keep going and still
# write the fixtures, which are what the app needs.
try:
    wb.save(OUT)
except PermissionError:
    print("WARNING: could not write", OUT, "(open in Excel?); fixtures still written")

templates = []
for (cid, name, discipline, allowed_roles, capacity, structure, options) in catalog:
    caps = [c.strip() for c in capacity.split(",")]
    reps = REPS.get(cid, [{"shape": SHAPE.get(cid, "simple")}])
    # REPS-defined exercises carry their own per-rep lung/speed; for the rest,
    # stamp a uniform lung volume / speed onto the placeholder rep when known.
    if cid not in REPS:
        for r in reps:
            if cid in LUNG:
                r["lung_volume"] = LUNG[cid]
            if cid in SPEED:
                r["pace"] = SPEED[cid]
    templates.append({
        "schema_version": 1,
        "id": cid,
        "name": name,
        "environment": ENV.get(cid, "pool"),
        "role": role_for(cid),
        "discipline": discipline,
        "capacity_tags": caps,
        "goal": structure,
        "cues": options,
        "log_mode": LOG_MODE.get(cid, "per_rep"),
        "set_repeat": SET_REPEAT.get(cid, 1),
        **({"recovery_inter": RECOVERY_INTER[cid]} if cid in RECOVERY_INTER else {}),
        "termination": TERMINATION.get(cid, {"type": "fixed_n"}),
        "reps": reps,
    })

# Convention guard: a fixed_n termination must carry no count. The planned rep
# count lives in reps.length x set_repeat, so a stray "n" here would reintroduce
# the two-sources-of-truth bug.
for t in templates:
    term = t["termination"]
    assert not (term.get("type") == "fixed_n" and "n" in term), (
        f"{t['id']}: fixed_n must not carry n (count = reps x set_repeat)"
    )

affinity_records = [
    {"exercises": [a, b], "weight": len(sessions), "sessions": sessions}
    for (a, b), sessions in affinities
]

fixtures = {"schema_version": 1, "templates": templates, "affinities": affinity_records}
with open(FIXTURES, "w", encoding="utf-8") as f:
    json.dump(fixtures, f, indent=2, ensure_ascii=False)
    f.write("\n")

# Seals edition: its own fixtures file with only the seals collection and no
# affinities, selected at app build time by VITE_EDITION=seals.
seals_fixtures = {"schema_version": 1, "templates": seals_templates, "affinities": []}
with open(FIXTURES_SEALS, "w", encoding="utf-8") as f:
    json.dump(seals_fixtures, f, indent=2, ensure_ascii=False)
    f.write("\n")

print("wrote", OUT)
print("wrote", FIXTURES)
print("wrote", FIXTURES_SEALS, "|", len(seals_templates), "seals templates")
print("catalog:", len(catalog), "| mapping rows:", len(mapping),
      "| affinity pairs:", len(affinities), "| templates:", len(templates))
